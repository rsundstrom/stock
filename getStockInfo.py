""" Gets date and data sets from the multpl web site and puts it into an excel spread sheet.
    The date is included for each column to verify the dates match across the row.
    Currently everything is working great. Use 'cleanData.py to expand everything to monthly data.

    Last edited by RayS 2/6/2018
"""

from lxml import html
import requests
import datetime as dt
from openpyxl import Workbook
import os


def scrape_data(url, remove=','):  # extract data and values from multpl.com site
    page = requests.get(url)  # grab html code for page
    tree = html.fromstring(page.text)  # extract the xml tree
    dates = tree.xpath('//td[@class="left"]/text()')  # look for class="left" to identify dates
    # print dates[:5]
    values = tree.xpath('//td[@class="right"]/text()')  # look for class="right" to identify values
    datesF = []  # import the date strings into datetime tuples
    for x in dates:
        # datesF.append(dt.datetime.strptime(x,'%b %d, %Y').date())
        datesF.append(x.strip())
    # print len(values)
    valuesF = []  # convert the value strings to formatted floats
    for x in values:
        try:
            # valuesF.append("{:8.2f}".format(float(x.strip().replace(remove,'')))) # makes a string
            valuesF.append((float(x.strip().replace(remove, ''))))  # makes a float
        except:
            # print'there was an invalid value!'
            pass
    return datesF, valuesF


SP_price = 'http://www.multpl.com/s-p-500-price/table?f=m'  # monthly to 1871
SP_PE = 'http://www.multpl.com/table?f=m'  # monthly to 1871
SP_div = 'http://www.multpl.com/s-p-500-dividend-yield/table?f=m'  # monthly to 1871, missing last quarter
SP_earnings = 'http://www.multpl.com/s-p-500-earnings/table?f=m'  # monthly to 1871, missing last half
US_Tres_10y = 'http://www.multpl.com/interest-rate/table?f=m'  # monthly, 1871+
US_inflation = 'http://www.multpl.com/inflation/table?f=m'  # monthly, 1872+, missing last quarter
Shiller_PE = 'http://www.multpl.com/shiller-pe/table?f=m'  # monthly 1881+
US_Home_Price = 'http://www.multpl.com/case-shiller-home-price-index-inflation-adjusted/table?f=m'  # mixed, 1890+, -last H
US_unemployment = 'http://www.multpl.com/unemployment/table?f=m'  # monthly 1948+, missing last month
US_LT_unemployment = 'http://www.multpl.com/us-long-term-unemployment-rate/table/by-month'  # monthly 1948+, missing last month
US_GDP_growth = 'http://www.multpl.com/us-gdp-growth-rate/table/by-quarter'  # quarterly, 1933+, missing last quarter
SP_PriceToSales = 'http://www.multpl.com/s-p-500-price-to-sales/table/by-quarter'  # quarterly to 2000

params = [[SP_price, ','], [SP_PE, ''], [SP_div, '%'], [SP_earnings, ''], [US_Tres_10y, '%'], [US_inflation, '%'],
          [Shiller_PE, ''], [US_Home_Price, ','], [US_unemployment, '%'], [US_LT_unemployment, '%'],
          [US_GDP_growth, '%']]
headings = [['SP_price', ','], ['SP_PE', ''], ['SP_div', '%'], ['SP_earnings', ''], ['US_Tres_10y', '%'],
            ['US_inflation', '%'],
            ['Shiller_PE', ''], ['US_Home_Price', ','], ['US_unemployment', '%'], ['US_LT_unemployment', '%'],
            ['US_GDP_growth', '%']]

data = []
x = 0
while x < len(params):
    # pass the URL and characters to remove like commas in floats, percent signs, etc
    data.append(scrape_data(params[x][0], remove=params[x][1]))
    if len(data[x][0]) != len(data[x][1]) or len(data[x][0]) == 0 or len(data[x][1]) == 0:
        print('Reading Error!', len(data[x][0]), len(data[x][1]))
    x += 1

print('Number of categories read =', len(data))

print("Path at terminal when executing this file")
print(os.getcwd() + "\n")

# os.chdir('/home/ray/Documents/programming/python/stocks/web_scraping')  # windows directory
filename = 'data/scrape_' + str(dt.date.today()) + '.xlsx'
print("filename = " + filename)

wb = Workbook()
ws = wb.worksheets[0]
ws.title = 'data'

#  Write the data array to the spreadsheet
col = 0
while col < len(data):
    ws.cell(row=1, column=2 * col + 1).value = "Date"
    ws.cell(row=1, column=2 * col + 2).value = headings[col][0]

    row = 0
    while row < min(len(data[col][0]), 5000):
        ws.cell(row=row + 2, column=2 * col + 1).value = data[col][0][row]
        ws.cell(row=row + 2, column=2 * col + 2).value = data[col][1][row]
        row += 1
    print('wrote column:', col, ', length =', len(data[col][0]))
    col += 1  # increment ahead 2 columns for the next date and data section

wb.save(filename)
