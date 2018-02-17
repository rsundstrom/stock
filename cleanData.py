""" This program cleans up the web scrape file by filling in the monthly data for quarterly
and annual data. It currently works for a all data collected by "getStockInfo.py".

Last edited by RayS 2/8/2018
"""

import datetime as dt
import calendar as cal
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

date_format = "%b %#d, %Y"  # "%b %#d, %Y" for windows, "%b %-d, %Y" for linux

def copy_in_columns(my_sheet):
    """ Gets the rows from a spreadsheet and then transposes them so the array index is in columns """
    my_list = []
    for row in my_sheet.iter_rows():
        temp_list = []
        for my_cell in row:
            temp_list.append(my_cell.value)
        my_list.append(temp_list)
    col_list = list(map(list, zip(*my_list)))  # transpose array to make it by column instead of row
    return col_list


def addMonths(sourcedate, months):
    """ adds (or subtracts) a month to a date """
    month = sourcedate.month - 1 + months
    year = int(sourcedate.year + month / 12 )
    month = month % 12 + 1
    day = min(sourcedate.day, cal.monthrange(year, month)[1])
    return dt.date(year, month, day)


def monthStart(date):
    """ converts last day of month to first day of next month """
    if date.day > 15:
        return date + dt.timedelta(hours=24)
    else:
        return date


def fill_in_data(data, col):
    # interpolate quarterly and annual data to fill in all months.
    print("length of arrays: ", len(data[col]), len(data[col+1]), "col = ", col)
    # create empty arrays for interpolated date and data
    newDate = []
    newData = []
    # compare date objects and look for difference grater than 1 month.
    x = 2
    print('First row data', data[col][x])
    while x < (min(len(data[col]), len(data[col+1])))-1:  # min() in case the date and data cols are not the same length
        if data[col][x+1] == "" or data[col][x+1] is None:
            x += 1
            continue  # skip is cell is empty or you are at the end of the column
        else:
            # convert two consecutive date strings to datetime objects
            # print(data[col][x])
            try:
                date1 = dt.datetime.strptime(data[col][x], '%b %d, %Y').date()
                date2 = dt.datetime.strptime(data[col][x+1], '%b %d, %Y').date()
            except Exception as e:
                print('EXCEPTION: Conversion to datetime failed: col =', col, ', row =', x)
                print("    ", e)
            # if date is end of month, convert to beginning of next month
            date1 = (monthStart(date1))
            date2 = (monthStart(date2))
            # find months difference in the two dates
            monDiff = (date1.year - date2.year) * 12 + (date1.month - date2.month)
            if monDiff < 1:
                print("Less than 1 month difference: col =", col, "row =", x+1, "dates:", date1, date2)
            # find the incremental change in each interpolated point
            try:
                dData = (data[col+1][x] - data[col+1][x+1])/monDiff
            except Exception as e:
                print("EXCEPTION: divide by 0, Line " + str(x))
                print("    ", e)
            # store beginning value for interpolated data
            data1 = data[col+1][x]
            data2 = data[col+1][x+1]  # save last data point, needed later
            # start lists with first actual date and data point
            newDate.append(date1.strftime(date_format))  # "%b %-d, %Y"
            newData.append(data1)

            # add interpolated monthly date and data points into the lists
            y = 0
            while y < monDiff -1:
                date1 = addMonths(date1, -1)  # add 1 month to data on each iteration
                # newDate.append(str(date1))
                newDate.append(date1.strftime(date_format))
                data1 = data1 - dData  # calculate interpolated data point
                newData.append(data1)
                y += 1
        x += 1
    newDate.append(date2.strftime(date_format))
    newData.append(data2)

    # fill in missing dates and data at the top of the column
    # find where the first date is located in reference date list
    datePlace = data[0].index(newDate[0])
    print("Starting date position: " + str(datePlace))
    z = 0
    while z < datePlace:
        newDate.insert(z, data[0][z])
        newData.insert(z, data[col+1][z])
        z += 1
    # print("length of data set after beginning correction = ", len(newDate))

    # replace the original list with the newly created list

    print("length of interpolated date and data = ", len(newDate), len(newData))
    # print('newDate', len(newDate), newDate)
    # print('newData', len(newData), newData)
    return newDate, newData


def write_to_file(data_l, out_file, all_dates=False):
    """ Write data to a new file. all_dates reprints the dates for each vector for debugging"""
    wbo = Workbook()
    wso = wbo.worksheets[0]
    wso.title = "clean_data"

    # write all lists to an excel file
    my_col = 1
    for c, col_o in enumerate(data_l):
        if c == 0 or c % 2 != 0 or all_dates:
            for r, row_o in enumerate(col_o):
                wso.cell(column=my_col, row=r+1).value = data_l[c][r]
            if c % 2 == 0:  # on even columns, set width to 12 (dates)
                wso.column_dimensions[get_column_letter(my_col)].width = 12
            my_col += 1
    wbo.save(out_file)


# os.chdir('/home/ray/Documents/programming/python/stocks/web_scraping') # set working directory
filename = 'scrape_' + str(dt.date.today()) + '.xlsx'
# filename = 'scrape_2018-02-06.xlsx'
sheetName = "data"
print("filename = " + filename)

wb = load_workbook(filename)
ws = wb.get_sheet_by_name(sheetName)

data_list = copy_in_columns(ws)  # data[row,column]
# print('data', data[:10])

print('shortest list', len(min(data_list, key=len)))
print('shortest list2', len(min(data_list, key=lambda coll: len(coll))))
print('shortest list3', min([len(ls) for ls in data_list]))

for ls in data_list:
    print("array length", len(ls))

# colMax = 22
# send each date/data pair to 'fill_in_data()' to interpolate the missing months
cols = 2  # S&P price is the reference so don't do the first column date/data pair
while cols < len(data_list):
    filled = fill_in_data(data_list, cols)
    data_list[cols] = filled[0]
    data_list[cols + 1] = filled[1]
    cols += 2

# output to a new spreadsheet
# output_file = 'clean_' + filename
output_file = 'clean_' + str(dt.date.today()) + '.xlsx'

# If 'all_dates' is true, the date for each vector is written to the file for debug.
write_to_file(data_list, output_file, all_dates=False)
